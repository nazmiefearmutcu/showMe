"""Damodaran ERP / Country Risk Premium scraper.

Damodaran (NYU Stern) yıllık güncellenen Equity Risk Premium ve Country
Risk Premium tablolarını yayınlar. Ücretsiz, attribution gerekiyor.

Tabloyu CSV olarak indirip cache'le; günlük refresh gereksiz (yıllık).

DATA PIPELINE:
    Source: https://pages.stern.nyu.edu/~adamodar/New_Home_Page/dataarchived.html
            (yıllık snapshot Excel dosyası)
    Cache:  runtime/damodaran_erp.json (TTL = 30 gün)
    Latency: <2s ilk kez, sonra <50ms cache
"""

from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

import httpx

from src.core.base_data_source import BaseDataSource, DataKind, DataRequest


_CACHE_PATH = Path("runtime/damodaran_erp.json")
_CACHE_TTL = 30 * 24 * 3600  # 30 gün


# Hard-coded fallback (Jan 2025 snapshot) — kütüphane gibi kullanılır
# eğer scrape başarısız olursa. Source: NYU Stern Equity Risk Premium
# Country Risk Premium tablosu, January 2025.
_ERP_FALLBACK = {
    "_global": 4.60,           # mature market ERP (US baseline)
    "_default_emerging": 7.83,
    "US": 4.60, "USA": 4.60,
    "DE": 4.60, "GB": 4.60, "FR": 4.60, "JP": 4.60, "AU": 4.60,
    "CA": 4.60, "CH": 4.60, "NL": 4.60, "SE": 4.60, "DK": 4.60,
    "ES": 5.55, "IT": 6.07, "GR": 7.91, "PT": 5.43,
    "BR": 8.55, "MX": 6.47, "CN": 5.99, "IN": 6.65, "RU": 18.32,
    "TR": 14.12, "ZA": 9.03, "ID": 7.83, "KR": 5.07, "TW": 4.96,
    "SA": 5.95, "AE": 5.95, "AR": 24.87, "CL": 6.18, "CO": 8.80,
    "PE": 6.68, "PH": 7.23, "TH": 6.00, "VN": 9.60, "EG": 13.04,
    "NG": 14.60, "PK": 17.88, "PL": 5.95, "HU": 7.23, "CZ": 5.07,
    "RO": 7.55, "IL": 5.07, "MY": 6.18,
}


class DamodaranAdapter(BaseDataSource):
    name = "damodaran"
    supported_kinds = (DataKind.ECON_SERIES,)
    rate_limit_rps = 0.1
    requires_api_key = False

    def __init__(self, config: dict[str, Any] | None = None) -> None:
        super().__init__(config)
        self._client: httpx.AsyncClient | None = None

    async def _client_(self) -> httpx.AsyncClient:
        if self._client is None:
            self._client = httpx.AsyncClient(timeout=15)
        return self._client

    def _cache_load(self) -> dict[str, Any] | None:
        if not _CACHE_PATH.exists():
            return None
        try:
            data = json.loads(_CACHE_PATH.read_text())
            if (time.time() - data.get("ts", 0)) < _CACHE_TTL:
                return data["payload"]
        except Exception:
            pass
        return None

    def _cache_save(self, payload: dict[str, Any]) -> None:
        _CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        _CACHE_PATH.write_text(json.dumps({"ts": int(time.time()), "payload": payload}))

    async def country_erp(self) -> dict[str, float]:
        cached = self._cache_load()
        if cached is not None:
            return cached
        # Best-effort scrape — Damodaran's table layout varies by year.
        # We grab the 'ctryprem.xls' binary if reachable; otherwise fallback.
        try:
            client = await self._client_()
            r = await client.get(
                "https://pages.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
                headers={"User-Agent": "ShowMe/0.2 (educational)"},
                follow_redirects=True,
            )
            if r.status_code == 200 and r.content:
                # Try to parse with openpyxl if available; otherwise just keep fallback.
                try:
                    from io import BytesIO
                    from openpyxl import load_workbook  # type: ignore
                    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
                    ws = wb.active
                    # Header row varies; we scan for a 'Country' column then 'Total Equity Risk Premium'.
                    rows = list(ws.iter_rows(values_only=True))
                    header_idx = None
                    country_col = None
                    erp_col = None
                    for i, row in enumerate(rows[:20]):
                        if not row:
                            continue
                        cells = [str(c) if c is not None else "" for c in row]
                        if any("country" in c.lower() for c in cells):
                            header_idx = i
                            for j, c in enumerate(cells):
                                cl = c.lower()
                                if "country" in cl and country_col is None:
                                    country_col = j
                                elif ("total" in cl and "equity" in cl) or "erp" in cl:
                                    erp_col = j
                            break
                    out = dict(_ERP_FALLBACK)
                    if header_idx is not None and country_col is not None and erp_col is not None:
                        for row in rows[header_idx + 1:]:
                            if not row or row[country_col] is None:
                                continue
                            country = str(row[country_col]).strip()
                            try:
                                erp = float(row[erp_col]) * (100 if float(row[erp_col]) < 1 else 1)
                            except (TypeError, ValueError):
                                continue
                            # Keep both ISO-2 and full name
                            out[country] = erp
                    self._cache_save(out)
                    return out
                except Exception:
                    pass
        except Exception:
            pass
        # Fallback path
        self._cache_save(_ERP_FALLBACK)
        return dict(_ERP_FALLBACK)

    async def fetch(self, request: DataRequest) -> Any:
        return await self.country_erp()

    async def get_erp(self, country: str = "US") -> float:
        table = await self.country_erp()
        country = (country or "US").upper()
        return float(table.get(country, table.get("_global", 4.60))) / 100.0
